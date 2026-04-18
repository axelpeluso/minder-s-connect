"""Seed the products table from bellezza_miami_inventory.xlsx.

Reads the 'Full Inventory' sheet, maps each row to a product, embeds
name + description via OpenAI text-embedding-3-small (1536-dim), and
upserts into the products table. Idempotent: row IDs are derived from
the SKU via uuid5, so re-running upserts in place.

Env required:
  DATABASE_URL    Supabase Postgres URL (session pooler).
  OPENAI_API_KEY

Install:
  pip install "psycopg[binary]" pgvector openai openpyxl
"""
from __future__ import annotations

import os
import re
import sys
import uuid
from pathlib import Path

import psycopg
from openai import OpenAI
from openpyxl import load_workbook
from pgvector.psycopg import register_vector

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "bellezza_miami_inventory.xlsx"
EMBED_MODEL = "text-embedding-3-small"
EMBED_DIM = 1536
NAMESPACE = uuid.UUID("8a4b1c3d-2e6f-4b8a-9c0d-1e2f3a4b5c6d")

CATEGORY_MAP = {
    "DIY Kit": "diy_kit",
    "Soft Gel Tips": "soft_gel_tips",
    "Cuticle Oil": "cuticle_care",
    "Nail Preparation": "nail_preparation",
}

EM_DASH = "\u2014"


def deterministic_id(sku: str) -> uuid.UUID:
    return uuid.uuid5(NAMESPACE, sku)


def slugify(s: str) -> str:
    s = s.strip().lower().replace("&", "and")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


def is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and value.strip() in ("", EM_DASH))


def build_sku_name_tags(
    category: str,
    name: str,
    shape: str | None,
    variant: str | None,
    notes: str,
) -> tuple[str, str, list[str]]:
    notes_l = notes.lower()
    name_l = name.lower()

    if category == "diy_kit":
        return (
            f"bz-diy-{shape}",
            name,
            ["hema_free", "beginner", "all_in_one", "no_lamp"],
        )

    if category == "soft_gel_tips":
        length_slug = slugify(variant or "default")
        sku = f"bz-tips-{shape}-{length_slug}"
        display = f"{name} {EM_DASH} {variant}" if variant else name
        return sku, display, [f"length:{length_slug}"]

    if category == "cuticle_care":
        scent_slug = slugify(variant or "")
        sku = f"bz-cuticle-{scent_slug}" if scent_slug else f"bz-cuticle-{slugify(name)}"
        tags = ["fast_absorbing", "non_greasy"]
        if scent_slug:
            tags.append(f"scent:{scent_slug}")
        return sku, name, tags

    if category == "nail_preparation":
        tags: list[str] = []
        if "toxic-free" in name_l or "toxic-free" in notes_l:
            tags.append("toxic_free")
        if "formaldehyde-free" in notes_l:
            tags.append("formaldehyde_free")
        if "waterproof" in notes_l:
            tags.append("waterproof")
        if "precision tip" in notes_l:
            tags.append("precision_tip")
        return "bz-glue-toxic-free", name, tags

    return slugify(name), name, []


def parse_xlsx() -> list[dict]:
    wb = load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Full Inventory"]
    products = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        idx, name, category, shape, variant_raw, notes, price, qty, _, status = row
        if not isinstance(idx, int) or not name or not category:
            continue

        category_key = CATEGORY_MAP.get(category)
        if not category_key:
            print(f"  skip (unknown category {category!r}): {name}", file=sys.stderr)
            continue

        shape_norm = None if is_blank(shape) else shape.strip().lower()
        variant = None if is_blank(variant_raw) else variant_raw.strip()
        notes_str = notes.strip() if isinstance(notes, str) else ""

        sku, display_name, tags = build_sku_name_tags(
            category_key, name.strip(), shape_norm, variant, notes_str,
        )

        desc_parts = []
        if notes_str:
            desc_parts.append(notes_str)
        if category_key == "soft_gel_tips" and variant:
            desc_parts.append(f"{variant} length, {shape_norm} shape.")
        elif category_key == "cuticle_care" and variant:
            desc_parts.append(f"{variant} scent.")
        if category_key == "diy_kit":
            desc_parts.append("No lamp needed. Lasts up to 3 weeks.")
        if category_key == "soft_gel_tips":
            desc_parts.append("Requires Nail Glue to apply.")

        products.append({
            "sku": sku,
            "name": display_name,
            "category": category_key,
            "shape": shape_norm,
            "tags": tags,
            "price_cents": int(round(float(price) * 100)),
            "in_stock": (status == "Active") and (qty or 0) > 0,
            "description": " ".join(desc_parts),
        })

    return products


def main() -> int:
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        print("DATABASE_URL not set", file=sys.stderr)
        return 1
    if not os.environ.get("OPENAI_API_KEY"):
        print("OPENAI_API_KEY not set", file=sys.stderr)
        return 1

    products = parse_xlsx()
    if not products:
        print("no products parsed", file=sys.stderr)
        return 1
    print(f"parsed {len(products)} products from xlsx")

    client = OpenAI()
    inputs = [
        f"{p['name']}. Category: {p['category']}. {p['description']} "
        f"Tags: {', '.join(p['tags'])}."
        for p in products
    ]
    resp = client.embeddings.create(model=EMBED_MODEL, input=inputs)
    embeddings = [d.embedding for d in resp.data]
    if any(len(v) != EMBED_DIM for v in embeddings):
        print(f"embedding dim mismatch (expected {EMBED_DIM})", file=sys.stderr)
        return 1

    rows = [
        (
            str(deterministic_id(p["sku"])),
            p["sku"],
            p["name"],
            p["category"],
            p["shape"],
            p["tags"],
            p["price_cents"],
            p["in_stock"],
            p["description"],
            v,
        )
        for p, v in zip(products, embeddings)
    ]

    with psycopg.connect(db_url) as conn:
        register_vector(conn)
        with conn.cursor() as cur:
            cur.executemany(
                """
                insert into products
                    (id, sku, name, category, shape, tags, price_cents,
                     in_stock, description, embedding)
                values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                on conflict (id) do update set
                    sku         = excluded.sku,
                    name        = excluded.name,
                    category    = excluded.category,
                    shape       = excluded.shape,
                    tags        = excluded.tags,
                    price_cents = excluded.price_cents,
                    in_stock    = excluded.in_stock,
                    description = excluded.description,
                    embedding   = excluded.embedding;
                """,
                rows,
            )
        conn.commit()

    print(f"upserted {len(rows)} products")
    return 0


if __name__ == "__main__":
    sys.exit(main())
